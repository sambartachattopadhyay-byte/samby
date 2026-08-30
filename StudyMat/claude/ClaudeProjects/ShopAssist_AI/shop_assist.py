from __future__ import annotations

import json
import os
from datetime import datetime
from typing import Any, Tuple

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import Workbook, load_workbook

load_dotenv()


class ShopAssistAI:

    # ── Connection ────────────────────────────────────────────────────────────

    def __init__(self):
        base_url = os.getenv("ANTHROPIC_BASE_URL", "").rstrip("/") + "/chat-models/"
        self.client = OpenAI(
            api_key=os.getenv("ANTHROPIC_API_KEY"),
            base_url=base_url,
            timeout=1200,
        )
        self.model = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-6")
        self.excel_path = os.getenv(
            "EXCEL_PATH",
            "/Users/samby/StudyMat/claude/ClaudeProjects/ShopAssist_AI/Data/prompt_history.xlsx",
        )
        self.sheet_name = os.getenv("SHEET_NAME", "PromptHistory")
        self.headers = ["Sequence", "Role", "Prompt", "TimeOfExecution"]
        self.system_prompt = (
            'start the response with  Samby AI  a helpful assistant for your online store. '
            'I will answer questions about orders, returns, and products."Only upto the text under quotes. '
            'Below are the instructions for you to follow when responding to customer queries. Be precise, '
            'to the point and professional and crispier try to respond within 20 words. Be Helpful and Polite '
            'but do not over commit to anything. If you don\'t know the answer, ask for more information.'
        )
        self.messages: list = []

    # ── Excel Definition ──────────────────────────────────────────────────────

    def create_excel(self) -> str:
        os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            ws.append(self.headers)
            wb.save(self.excel_path)
        return self.excel_path

    def _load_ws(self):
        wb = load_workbook(self.excel_path)
        ws = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.active
        return wb, ws

    def _next_sequence(self, ws) -> int:
        highest = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            try:
                highest = max(highest, int(row[0]))
            except (TypeError, ValueError):
                continue
        return highest + 1

    def add_prompt(self, role: str, prompt: Any) -> int:
        self.create_excel()
        wb, ws = self._load_ws()
        sequence = self._next_sequence(ws)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        prompt_text = prompt if isinstance(prompt, str) else str(prompt)
        ws.append([sequence, role, prompt_text, timestamp])
        wb.save(self.excel_path)
        wb.close()
        return sequence

    def clear_content(self) -> str:
        if not os.path.exists(self.excel_path):
            self.create_excel()
            return self.excel_path
        wb, ws = self._load_ws()
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        wb.save(self.excel_path)
        wb.close()
        return self.excel_path

    # ── Test Evaluation ───────────────────────────────────────────────────────

    def classify_intent(self, customer_message: str) -> dict:
        prompt = f"""
Classify the customer's message into one of these intents:
- refund_request
- return_order
- track_shipment
- order_status
- billing_issue
- product_question
- other

Customer message:
{customer_message}

Return only a valid JSON object.
Do not include markdown.
Do not include explanations.
Do not wrap the JSON in a code block.
{{
  "intent": "refund_request"
}}
"""
        response = self.client.chat.completions.create(
            model=self.model,
            max_tokens=200,
            temperature=0,
            messages=[{"role": "user", "content": prompt}],
        )
        return json.loads(response.choices[0].message.content)

    def run_intent_tests(self, test_cases: list[dict]) -> None:
        for test_case in test_cases:
            result = self.classify_intent(test_case["input"])
            actual = result.get("intent")
            expected = test_case["expected_intent"]
            if actual == expected:
                print(f"PASS | {test_case['input']}")
            else:
                print(f"FAIL | {test_case['input']} | expected={expected}, got={actual}")

    # ── Chat Execution ────────────────────────────────────────────────────────

    def add_user_message(self, text: str) -> list:
        self.messages.append({"role": "user", "content": text})
        self.add_prompt("user", text)
        return self.messages

    def add_assistant_message(self, text: str) -> list:
        self.messages.append({"role": "assistant", "content": text})
        self.add_prompt("assistant", text)
        return self.messages

    def add_system_message(self, text: str) -> list:
        self.messages.append({"role": "system", "content": text})
        self.add_prompt("system", text)
        return self.messages

    def get_chat_response(self) -> Tuple[str, str]:
        response = self.client.chat.completions.create(
            model=self.model,
            temperature=0.7,
            max_tokens=300,
            messages=self.messages,
        )
        return response.choices[0].message.content, response.choices[0].finish_reason

    def chat(self, user_text: str) -> str:
        self.add_user_message(user_text)
        reply, _ = self.get_chat_response()
        self.add_assistant_message(reply)
        return reply

    def reset(self) -> None:
        self.messages = []
        self.clear_content()
        self.add_system_message(self.system_prompt)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    assistant = ShopAssistAI()

    test_cases = [
        {"input": "I want to return an order",              "expected_intent": "return_order"},
        {"input": "Where is my order? I want to track it",  "expected_intent": "track_shipment"},
        {"input": "No idea where is the order, I want to cancel it", "expected_intent": "return_order"},
        {"input": "I was charged twice for the same order", "expected_intent": "billing_issue"},
    ]

    print("── Intent Tests ──")
    assistant.run_intent_tests(test_cases)

    print("\n── Chat Demo ──")
    assistant.reset()
    reply = assistant.chat("I have ordered a headphone last week, it does not work, I want my money back !!!")
    print(f"Assistant: {reply}")
    reply = assistant.chat("My order number is 12345")
    print(f"Assistant: {reply}")

    print("\nFull message history:")
    for msg in assistant.messages:
        print(f"  [{msg['role']}] {msg['content'][:80]}")


if __name__ == "__main__":
    main()
