"""Prompt history logger — stores every prompt exchanged with the assistant
in an Excel workbook.

Columns: Sequence | Role | Prompt | TimeOfExecution
  - Sequence      : auto-generated integer, increments with every new row
  - Role          : caller-supplied ("system" / "user" / "assistant" / ...)
  - Prompt        : caller-supplied text (or dict/list — coerced to str)
  - TimeOfExecution: auto-stamped with datetime.now()
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook

DATA_DIR = "/Users/samby/StudyMat/claude/ClaudeProjects/ShopAssist_AI/Data"
EXCEL_PATH = os.path.join(DATA_DIR, "prompt_history.xlsx")
SHEET_NAME = "PromptHistory"
HEADERS = ["Sequence", "Role", "Prompt", "TimeOfExecution"]


def create_excel(file_path: str = EXCEL_PATH) -> str:
    """Create the workbook with the header row if it does not already exist."""
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
        wb.save(file_path)
    return file_path


def _load_ws(file_path: str):
    wb = load_workbook(file_path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    return wb, ws


def _next_sequence(ws) -> int:
    highest = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            highest = max(highest, int(row[0]))
        except (TypeError, ValueError):
            continue
    return highest + 1


def add_prompt(role: str, prompt: Any, file_path: str = EXCEL_PATH) -> int:
    """Append one row for (role, prompt). Auto-fills Sequence and TimeOfExecution.

    `prompt` may be a string, dict, or list (e.g. the OpenAI-style messages list) —
    non-strings are stringified so the whole payload can be captured.
    Returns the sequence number assigned to the new row.
    """
    create_excel(file_path)
    wb, ws = _load_ws(file_path)
    sequence = _next_sequence(ws)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    prompt_text = prompt if isinstance(prompt, str) else str(prompt)
    ws.append([sequence, role, prompt_text, timestamp])
    wb.save(file_path)
    wb.close()
    return sequence


def clear_content(file_path: str = EXCEL_PATH) -> str:
    """Remove all data rows but keep the header."""
    if not os.path.exists(file_path):
        create_excel(file_path)
        return file_path
    wb, ws = _load_ws(file_path)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    wb.save(file_path)
    wb.close()
    return file_path


if __name__ == "__main__":
    create_excel()

    sample_messages = [
        {"role": "system", "content": "Samby AI a helpful assistant for your online store."},
        {"role": "user", "content": "I have ordered a headphone last week, it does not work, I want my money back!"},
        {"role": "assistant", "content": "I'm sorry to hear that. Could you share your order number?"},
        {"role": "user", "content": "My order number is 12345"},
    ]

    for msg in sample_messages:
        seq = add_prompt(msg["role"], msg["content"])
        print(f"logged  sequence={seq}  role={msg['role']}")

    print(f"\nWorkbook: {EXCEL_PATH}")
